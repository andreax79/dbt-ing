#!/usr/bin/env python

import re
import os
import os.path
import json
import yaml
import click
import shutil
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook  # type: ignore
from openpyxl.workbook import Workbook  # type: ignore
from .utils import (
    to_bool,
    Columns,
    Config,
    Table,
    Tables,
    FORMAT_JSON,
    DEFAULT_FIELD_DELIMITER,
    DEFAULT_SOURCE_FORMAT,
    DEFAULT_PARTITIONS_STYLE,
    PARTITIONS_STYLE_HIVE,
    PARTITIONS_STYLE_NON_HIVE,
)

__all__ = [
    "xlsx_to_json_mapping",
    "MappingException",
]


def represent_ordereddict(dumper, data):  # type: ignore
    value = []
    for item_key, item_value in data.items():
        node_key = dumper.represent_data(item_key)
        node_value = dumper.represent_data(item_value)
        value.append((node_key, node_value))
    return yaml.nodes.MappingNode("tag:yaml.org,2002:map", value)


yaml.add_representer(dict, represent_ordereddict)


class quoted(str):
    pass


# define a custom representer for strings
def quoted_presenter(dumper, data):  # type: ignore
    return dumper.represent_scalar("tag:yaml.org,2002:str", data, style='"')


yaml.add_representer(quoted, quoted_presenter)


def filename_to_source_table(filename: str, flow: str) -> str:
    "Convert filename into source table name"
    t = re.sub("\..*$", "", filename)  # remove extension
    t = re.sub("{{[^}]*}}", "", t).replace("-", "_").strip("_ . -")  # remove jinja macro
    t = t.replace("*", "").replace("?", "").strip("_")
    t = t.lower()
    if not t.startswith(flow + "_"):
        t = flow + "_" + t
    return t


def get_target_table_name(target_table: str, flow: str) -> str:
    result = target_table.lower()
    if not result.startswith(flow + "_"):
        result = flow + "_" + result
    return result


def table_to_path(table: str, flow: str) -> str:
    return flow + "/" + table[len(flow) + 1 :]


class MappingException(Exception):
    pass


class Empty(Exception):
    pass


def parse_row_config(table_def: Table, columns: Columns, row: List[Any]) -> None:
    try:
        if not row[1]:
            value = ""
        elif isinstance(row[1], str):
            value = row[1].strip()
        else:
            value = row[1]
        table_def[row[0].strip().lower()] = value
    except Exception as ex:
        raise MappingException("Error parsing row {}: {}".format(row, ex))


def parse_row_where(table_def: Table, columns: Columns, row: List[Any]) -> None:
    where_condition = (row[0] or "").strip()
    if where_condition:
        table_def["where_condition"] = where_condition


def parse_row_select_header(table_def: Table, columns: Columns, row: List[Any]) -> None:
    table_def["header"] = dict((h.lower(), i) for i, h in enumerate(row) if h is not None)
    table_def["parse_row"] = parse_row_select


def parse_row_select(table_def: Table, columns: Columns, row: List[Any]) -> None:
    item: Dict[str, Any] = {}
    for h, i in table_def["header"].items():
        try:
            value = row[i]
            if isinstance(value, str):
                value = value.strip()
        except Exception:
            value = None
        item[h] = value
    columns.append(item)


def parse_worsheet(ws: Workbook, flow: str, config: Config) -> Table:
    flow = flow.lower()
    columns: Columns = []
    if ws["A1"].value != "TYPE":
        click.secho("- skip worksheet {}".format(ws.title))
        return

    table_def: Table = {
        "parse_row": None,
        "type": ws["A2"].value,
        "flow": flow,
        "batch_profile": config["DBT__BATCH_PROFILE"],
        "datalake_profile": config["DBT__DATALAKE_PROFILE"],
        "batch_location": config["S3__BATCH_LOCATION"],
        "datalake_location": config["S3__DATALAKE_LOCATION"],
        "source_filename": None,
        "source_format": DEFAULT_SOURCE_FORMAT,
        "source_location": None,
        "source_schema": config["DB__SOURCE_SCHEMA"],
        "source_table": None,
        "target_location": None,
        "target_schema": config["DB__TARGET_SCHEMA"],
        "target_table": None,
        "field_delimiter": DEFAULT_FIELD_DELIMITER,
        "partitions": config["DEFAULT_PARTITIONS"],
        "partitions_style": DEFAULT_PARTITIONS_STYLE,
        "columns": columns,
    }
    if table_def["type"] != "#INGESTION":
        raise MappingException("#{} invalid type {}".format(0, table_def["type"]))
    for i, row in enumerate(ws.values, start=1):
        if i > 2:
            if row[0] == "#CONFIG":
                table_def["parse_row"] = parse_row_config
            elif row[0] == "#WHERE":
                table_def["parse_row"] = parse_row_where
            elif row[0] == "#SELECT":
                table_def["parse_row"] = parse_row_select_header
            else:
                table_def["parse_row"](table_def, columns, row)
    del table_def["parse_row"]
    del table_def["header"]
    # Source table
    if not table_def["source_table"]:
        try:
            table_def["source_table"] = filename_to_source_table(table_def["source_filename"], flow)
        except Exception:
            raise MappingException("invalid filename {}".format(table_def["source_filename"]))
    # Source location
    if not table_def["source_location"]:
        table_def["source_location"] = (
            os.path.join(
                table_def["batch_location"],
                table_to_path(table_def["source_table"], flow),
            )
            + "/"
        )
    # Target location
    if not table_def["target_location"]:
        table_def["target_location"] = (
            os.path.join(
                table_def["datalake_location"],
                table_to_path(table_def["target_table"], flow),
            )
            + "/"
        )
    return table_def


def parse_xlsx(
    filename: str,
    flow: str,
    config: Config,
) -> Tables:
    # Read xlsx
    flow = flow.lower()
    wb: Workbook = load_workbook(filename, read_only=True)
    try:
        tables: Tables = {}
        for ws in wb.worksheets:
            # Paese the worksheet
            table_def: Table = parse_worsheet(
                ws=ws,
                flow=flow,
                config=config,
            )
            # Insert the table in the tables
            target_table = get_target_table_name(table_def["target_table"], flow)
            click.secho("- worksheet {} table {}".format(ws.title, target_table))
            if target_table in tables:
                raise MappingException("Duplicated target table {}".format(target_table))
            tables[target_table] = table_def
        return tables
    finally:
        # Close the workbook after reading
        wb.close()


def parse_rows(tables: Tables, flow: str, config: Config) -> Tables:
    flow = flow.lower()
    for target_table, table_def in list(tables.items()):
        if to_bool(table_def.get("exclude_table")):
            click.secho("- skip target table %s" % target_table)
            del tables[target_table]
            continue
        click.secho("- %s" % target_table)
        columns = []
        table_def["col_number"] = 0

        for row_number, item in enumerate(table_def["columns"], start=1):
            try:
                if not item.get("source_name"):
                    raise Empty
                st = target_table
                col: Dict[str, Any] = {}
                source_format: str = table_def.get("source_format")  # type: ignore
                col_number = table_def["col_number"] + 1
                table_def["col_number"] = col_number
                table_def["columns"].append(col)
                if source_format == FORMAT_JSON:
                    col["source_column"] = item["source_name"].lower()
                else:
                    col["source_column"] = "col{}".format(col_number)
                    if to_bool(table_def.get("use_source_name")):
                        col["source_formula"] = item["source_name"]
                if not item.get("new_name"):
                    raise MappingException("#{} invalid target_column {}".format(row_number, "<empty>"))
                col["target_column"] = item["new_name"].lower()
                if "." in col["target_column"] or " " in col["target_column"] or "-" in col["target_column"]:
                    raise MappingException("#{} invalid target_column {}".format(row_number, col["target_column"]))
                col["description"] = item["description"] or ""
                col["data_type"] = item["data_type"].lower()
                col["key"] = to_bool(item["key"])
                col["index"] = to_bool(item["index"])
                col["nullable"] = to_bool(item["nullable"])
                if not col["data_type"]:
                    raise MappingException("#{} missing data_type in row {}".format(row_number, str(item)))
                elif col["data_type"] in ("char", "varchar"):
                    col["data_type"] = "{}({})".format(col["data_type"], item.get("length") or 255)
                elif col["data_type"] in ("decimal"):
                    col["data_type"] = "{}({})".format(
                        col["data_type"],
                        (str(item.get("length") or "18,2")).replace(".", ","),
                    )
                elif col["data_type"] in ("float"):
                    col["data_type"] = "real"
                elif col["data_type"] in ("double"):
                    col["data_type"] = "double precision"
                elif col["data_type"] in ("date", "timestamp"):
                    col["format"] = item.get("length")  # length column is used for format for the date/timestamp
                elif col["data_type"] in (
                    "boolean",
                    "tinyint",
                    "smallint",
                    "integer",
                    "bigint",
                    "float",
                    "double precision",
                ):
                    pass
                elif col["data_type"] in ("array<char>", "array<varchar>"):
                    col["data_type"] = "{}({})>".format(col["data_type"].rstrip(">"), item.get("length") or 255)
                else:
                    raise MappingException("#{} invalid data_type in row {}".format(row_number, str(item)))
                columns.append(col)
            except Empty:
                # click.secho('#{} empty line'.format(row_number), fg='red')
                pass

        # Add partition columns
        try:
            partitions = table_def["partitions"]
            if partitions == "":
                partitions = []
            elif isinstance(partitions, str):
                partitions = [x.strip() for x in table_def["partitions"].split(",")]
        except Exception:
            partitions = config["DEFAULT_PARTITIONS"]
        table_def["partitions"] = partitions
        for i, column in enumerate(partitions):
            if table_def["partitions_style"] == PARTITIONS_STYLE_NON_HIVE:
                source_column = "partition_{}".format(i)
            elif table_def["partitions_style"] == PARTITIONS_STYLE_HIVE:
                source_column = column
            else:
                raise MappingException("Invalid partitions style {}".format(table_def["partitions_style"]))
            if column == "year":
                data_type = "char(4)"
            elif column in ("day", "month", "hour", "minute"):
                data_type = "char(2)"
            else:
                data_type = "varchar(20)"
            column = {
                "source_column": source_column,
                "target_column": column,
                "description": "Partition {}".format(source_column),
                "partition": "yes",
                "data_type": data_type,
            }
            columns.append(column)
        table_def["columns"] = columns

    # Write json
    os.makedirs("flows", exist_ok=True)
    with open(os.path.join("flows", flow + ".cfg.json"), "w") as f:
        f.write(json.dumps(tables, indent=2))
    return tables


def generate_column(
    column: Dict[str, Any],
    kind: str,
    table_def: Table,
    forced_type: Optional[str] = None,
    has_compound_key: bool = False,
    table_format: Optional[str] = None,
) -> Dict[str, Any]:
    if kind == "batch" and not column.get("meta", {}).get("partition") and not table_format == FORMAT_JSON:
        forced_type = "varchar(65535)" if not column.get("meta", {}).get("partition") else None
    result = {
        "name": column["source_column"] if kind == "batch" else column["target_column"],
        "description": column["description"],
        "data_type": forced_type or column["data_type"],
        "tests": [],
        "meta": {},
    }
    if column.get("partition") == "yes":
        result["meta"]["partition"] = "yes"
    if kind == "datalake":
        if column.get("key") and not has_compound_key:
            result["tests"].append(
                {
                    "unique_where": {
                        "where": quoted(
                            " and ".join("{f}='{{{{ var('{f}') }}}}'".format(f=f) for f in table_def["partitions"])
                        )
                    }
                }
            )
    if not column.get("nullable") and column.get("partition") != "yes":
        result["tests"].append(
            {
                "not_null": {
                    "where": quoted(
                        " and ".join("{f}='{{{{ var('{f}') }}}}'".format(f=f) for f in table_def["partitions"])
                    )
                }
            }
        )
    if not result["meta"]:
        del result["meta"]
    if not result["tests"]:
        del result["tests"]
    return result


def generate_sources(tables: Tables, kind: str, flow: str) -> None:
    sources_path = "./models/sources/{kind}/{flow}".format(kind=kind, flow=flow)
    shutil.rmtree(sources_path, ignore_errors=True)
    os.makedirs(sources_path, exist_ok=True)

    for table_def in tables.values():
        if kind == "batch":
            name = table_def["batch_profile"] + "__" + table_def["source_table"]
            description = table_def.get("description") or "{flow} - {name}".format(flow=flow, name=name)
            source_format = table_def["source_format"]
            table_properties = {
                "CrawlerSchemaDeserializerVersion": "1.0",
                "CrawlerSchemaSerializerVersion": "1.0",
                "typeOfData": "file",
                "compressionType": table_def.get("compression_type", "none"),
            }
            if source_format == FORMAT_JSON:
                table_properties["classification"] = "json"
            else:
                table_properties["areColumnsQuoted"] = "false"
                table_properties["classification"] = "csv"
                table_properties["columnsOrdered"] = "true"
                table_properties["delimiter"] = table_def["field_delimiter"]
            doc = {
                "version": 2,
                "sources": [
                    {
                        "name": table_def["batch_profile"],
                        "tables": [
                            {
                                "name": name,
                                "description": description,
                                "identifier": table_def["source_table"],
                                "meta": {
                                    "flow": table_def["flow"],
                                    "field_delimiter": table_def["field_delimiter"],
                                    "file_format": "textfile",
                                    "location": table_def["source_location"],
                                    "format": source_format,
                                    "table_properties": table_properties,
                                },
                                "columns": [
                                    generate_column(
                                        column,
                                        kind,
                                        table_def,
                                        table_format=source_format,
                                    )
                                    for column in table_def["columns"]
                                ],
                            }
                        ],
                    }
                ],
            }
        else:
            name = table_def["datalake_profile"] + "__" + table_def["target_table"]
            key_columns = [column["target_column"] for column in table_def["columns"] if column.get("key")]
            has_compound_key = len(key_columns) > 1
            if not has_compound_key:
                tests = None
            else:
                tests = [
                    {
                        "unique_columns_where": {
                            "combination_of_columns": key_columns,
                            "where": quoted(
                                " and ".join("{f}='{{{{ var('{f}') }}}}'".format(f=f) for f in table_def["partitions"])
                            ),
                        }
                    }
                ]
            tags = table_def.get("tags").split() if table_def.get("tags") else None  # type: ignore
            doc = {
                "version": 2,
                "sources": [
                    {
                        "name": table_def["datalake_profile"],
                        "tables": [
                            {
                                "name": name,
                                "description": table_def.get("description")
                                or "{flow} - {name}".format(flow=flow, name=name),
                                "identifier": table_def["target_table"],
                                "meta": {
                                    "flow": table_def["flow"],
                                    "file_format": "parquet",
                                    "location": table_def["target_location"],
                                    "table_properties": {
                                        "classification": "parquet",
                                    },
                                },
                                "tests": tests,
                                "tags": tags,
                                "columns": [
                                    generate_column(
                                        column,
                                        kind,
                                        table_def,
                                        has_compound_key=has_compound_key,
                                    )
                                    for column in table_def["columns"]
                                ],
                            }
                        ],
                    }
                ],
            }
            for k in ["tests", "tags"]:
                if not doc["sources"][0]["tables"][0][k]:  # type: ignore
                    del doc["sources"][0]["tables"][0][k]  # type: ignore
        with open(os.path.join(sources_path, name + ".yml"), "w") as f:
            f.write(yaml.dump(doc, width=128))


def check(tables: Tables) -> None:
    has_duplicates = False
    for table in tables.values():
        cols = set()
        for column in table["columns"]:
            if column["target_column"] in cols:
                click.secho(
                    "Duplicate column '{column}' in table '{table}'".format(
                        column=column["target_column"], table=table["target_table"]
                    ),
                    fg="red",
                )
                has_duplicates = True
            cols.add(column["target_column"])
    if has_duplicates:
        raise MappingException("Duplicated columns")


def xlsx_to_json_mapping(
    filename: str,
    flow: str,
    config: Config,
) -> None:
    tables = parse_xlsx(
        filename=filename,
        flow=flow,
        config=config,
    )
    tables = parse_rows(tables, flow, config)
    check(tables)
    generate_sources(tables, "batch", flow)
    generate_sources(tables, "datalake", flow)
